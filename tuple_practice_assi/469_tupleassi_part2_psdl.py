import openpyxl
from openpyxl.styles import PatternFill

# Load the Excel file
path = "school_data.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# Get total rows and columns
row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows:", row)
print("Total Columns:", column)

# Print first column values
print("\nValue of first column:")
for i in range(1, row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    print(cell_obj.value)

# Print first row values
print("\nValue of first row:")
for i in range(1, column + 1):
    cell_obj = sheet_obj.cell(row=2, column=i)
    print(cell_obj.value, end=" ")
print()

# Get user input for a specific cell
cell_row = int(input("\nEnter the row number of the cell: "))
cell_col = int(input("Enter the column number of the cell: "))

# Print the selected cell value
cell_value = sheet_obj.cell(row=cell_row, column=cell_col)
print(f"Value at ({cell_row}, {cell_col}): {cell_value.value}")
# Change background color of the selected cell
cell = sheet_obj.cell(row=cell_row, column=cell_col)
cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
wb_obj.save("styled_school_data.xlsx")
print("\nCell color changed successfully!")

# Print entire Excel sheet
print("\nComplete Excel Data:")
for i in range(1, row + 1):
    for j in range(1, column + 1):
        cell_value = sheet_obj.cell(row=i, column=j).value
        print(cell_value, end=" ")
    print()

# Delete a row (Row 3)
sheet_obj.delete_rows(3)
wb_obj.save("updated_school_data.xlsx")
print("\nRow 3 deleted successfully!")

# Print the updated Excel sheet
print("\nUpdated Excel Data:")
for i in range(1, row):  # Adjusted loop to avoid empty row printing
    for j in range(1, column + 1):
        cell_value = sheet_obj.cell(row=i, column=j).value
        print(cell_value, end=" ")
    print()

# Search for a specific value in the Excel sheet
search_value = "Oakridge Grammar School"
found = False

for i in range(1, row + 1):
    for j in range(1, column + 1):
        if sheet_obj.cell(row=i, column=j).value == search_value:
            print(f"Value '{search_value}' found at ({i},{j})")
            found = True

if not found:
    print(f"Value '{search_value}' not found in the Excel file.")

'''
PS C:\Users\Shruti\OneDrive\Desktop\py assi sem 4\assi 4> python 469_assi4_part2_psdl.py
Total Rows: 11
Total Columns: 6

Value of first column:
School ID
101
102
103
104
105
106
107
108
109
110

Value of first row:
101 Greenfield Public School New York NY 1200 1985

Enter the row number of the cell: 2
Enter the column number of the cell: 4
Value at (2, 4): NY

Cell color changed successfully!

Complete Excel Data:
School ID School Name City State Total Students Established Year
101 Greenfield Public School New York NY 1200 1985
102 Riverdale High School Los Angeles CA 1500 1992
103 Springfield International Chicago IL 1100 2001
104 Maplewood Academy Houston TX 900 1978    
105 Bluebell Elementary Miami FL 800 1999    
106 Starlight High School Boston MA 1400 1988

107 Sunflower Public School San Diego CA 1000 2005
108 Oakridge Grammar School Denver CO 950 1995
109 Pinehill Secondary School Seattle WA 1100 2003
110 Redwood High School San Francisco CA 1300 1990

Row 3 deleted successfully!
104 Maplewood Academy Houston TX 900 1978
105 Bluebell Elementary Miami FL 800 1999
106 Starlight High School Boston MA 1400 1988
107 Sunflower Public School San Diego CA 1000 2005
108 Oakridge Grammar School Denver CO 950 1995
109 Pinehill Secondary School Seattle WA 1100 2003
110 Redwood High School San Francisco CA 1300 1990
Value 'Oakridge Grammar School' found at (8,2)
'''


